"""Perform reimbursement calculations and return output."""

import os
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from directors_reimbursements.common import Dates
from directors_reimbursements.config import config
from directors_reimbursements import logger

from directors_reimbursements.constants import (
    SHEET_NAME, INITIALS_COL, NAME_COL, EMAIL_COL, USERNAME_COL, MON_DATE_COL,
    WED_DATE_COL, ACTIVE_COL, DATE_FORMAT)

HEADING = ('Name', 'username', 'BBO$', 'Dates directed', 'Total dollars')


class Director():
    def __init__(self, initials, name, email, username, dates, active):
        self.initials = initials
        self.name = name
        self.email = email
        self.dates = dates
        self._dollars = 0
        self.first_name = self._get_first_name()
        self.username = username
        self.active = active

    def __repr__(self) -> str:
        return f'{self.initials} {self.name}'

    @property
    def dollars(self):
        # pylint: disable=no-member)
        return len(self.dates) * config.payment_bbo

    def _get_first_name(self):
        return self.name.split(' ')[0]


def calculate(dates: Dates) -> None:
    # pylint: disable=no-member)
    date_from = dates.start_date.strftime('%d %b %Y')
    date_to = dates.end_date.strftime('%d %b %Y')
    logger.info(f'Calculation started for {date_from} to {date_to}')
    workbook_path = Path(os.path.expanduser('~'), config.workbook_path)
    workbook = load_workbook(filename=workbook_path, data_only=True)

    directors = _get_directors(workbook)
    _get_dates_directed(dates, workbook, directors)
    csv_report = _create_csv_report(directors)
    formatted_report = _create_formatted_report(directors)

    return (directors, formatted_report, csv_report)


def _create_formatted_report(directors: dict[str, Director]) -> list[str]:
    (name, username, bbo_dollars, dates, total) = HEADING
    total_dollars = 0

    report = [(f'{name:<20} {username:<10} {bbo_dollars:>4} {dates}')]
    for director in directors.values():
        if director.active and director.dollars:
            report.append(
                (f'{director.name:<20} '
                 f'{director.username:<10} {director.dollars:>4} '
                 f'{", ".join(director.dates)}')
                 )
            total_dollars += director.dollars
    report.append(f'{total:<20} {"":<10} {total_dollars:>4}')
    logger.info("Created formatted report")
    return report


def _create_csv_report(directors: dict[str, Director]) -> list[str]:
    (name, username, bbo_dollars, dates, total) = HEADING
    total_dollars = 0

    report = [(f'{name},{username},{bbo_dollars},{dates}')]
    for director in directors.values():
        if director.active and director.dollars:
            report.append(
                (f'{director.name},'
                 f'{director.username}, {director.dollars},'
                 f'{", ".join(director.dates)}')
                 )
            total_dollars += director.dollars
    report.append(f'{total}, , {total_dollars}')

    logger.info("Created csv report")
    return report


def _get_dates_directed(
        dates: Dates,
        workbook: object,
        directors: dict[str, Director]) -> dict[str: str]:
    """Return a dict of directors and the dates they've directed."""
    worksheet = workbook[SHEET_NAME]
    start_date, end_date = dates.start_date, dates.end_date

    directed = {}
    for row in worksheet.iter_rows(values_only=True):
        if isinstance(row[0], datetime):
            for date_col in [MON_DATE_COL, WED_DATE_COL]:
                dir_col = date_col + 1
                alt_dir_col = date_col + 2
                if (row[dir_col] and row[date_col]
                        and start_date <= row[date_col] < end_date):
                    director = directors[row[dir_col]]
                    if row[alt_dir_col]:
                        director = directors[row[alt_dir_col]]
                    director.dates.append(row[date_col].strftime(DATE_FORMAT))
                    if row[dir_col] not in directed:
                        directed[row[dir_col]] = []
                    directed[row[dir_col]].append(
                        row[date_col].strftime(DATE_FORMAT))

    logger.info(f"Retrieved {len(directed)} directed date records")
    return directed


def _get_directors(workbook: object) -> dict[str, Director]:
    """Return a dict of Directors."""
    worksheet = workbook['Directors']
    directors = {}
    for row in worksheet.iter_rows(values_only=True):
        if row[0] and row[0] != 'Initials':
            director = Director(initials=row[INITIALS_COL],
                                name=row[NAME_COL],
                                email=row[EMAIL_COL],
                                username=row[USERNAME_COL],
                                dates=[],
                                active=row[ACTIVE_COL] is not None)
            directors[director.initials] = director
    logger.info(f"Retrieved {len(directors)} directors' records")
    return directors
